import sys
import os
import shutil
from openpyxl import load_workbook

MAX_LANG_NUM = 10

def readFile(path):
    with open(path, "rt") as f:
        text = f.read()
        return text

def writeFile(path, text):
    with open(path, "w") as f:
        text = f.write(text)

def copyLangFolders(langs, dst, src):
    for lang in langs:
        print(lang)
        dstFolder = dst + "/" + lang
        if os.path.exists(dstFolder):
            shutil.rmtree(dstFolder)
        shutil.copytree(src, dstFolder, 1024*1024)

def getLangs(wb):
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    langs = []
    for i in range(2, MAX_LANG_NUM):
        lang = sheet.cell(row=1, column=i).value
        if lang == None:
            break
        langs.append(lang)
    return langs
  

def fileTranslate(fileText, fileSheet, commonSheet, dst):
    originColNum = 0
    for i in range(1, MAX_LANG_NUM):
        if fileSheet.cell(row=1, column=i).value == "origin":
            originColNum = i
            break
    
    colNum = 1
    while fileSheet.cell(row=1, column=colNum).value != None:
        if colNum != originColNum:
            curLang = fileSheet.cell(row=1, column=colNum).value
            writePath = dst.format(lang=curLang)
            langText = fileText

            rowNum = 1
            word = commonSheet.cell(row=rowNum, column=colNum).value
            while word != None:
                originWord = commonSheet.cell(row=rowNum, column=originColNum).value
                langText = langText.replace(originWord, word)
                rowNum += 1
                word = commonSheet.cell(row=rowNum, column=colNum).value
            rowNum = 1
            word = fileSheet.cell(row=rowNum, column=colNum).value
            while word != None:
                originWord = fileSheet.cell(row=rowNum, column=originColNum).value
                langText = langText.replace(originWord, word)
                rowNum += 1
                word = fileSheet.cell(row=rowNum, column=colNum).value
            writeFile(writePath, langText)
        colNum += 1

def translateAll(wb, src, dst):
    sheetsNames = wb.sheetnames
    langs = getLangs(wb)
    copyLangFolders(langs, dst, src)
    commonSheet = None

    for shName in sheetsNames:
        if shName == "common":
            commonSheet = wb[shName]
            break

    for shName in sheetsNames:
        if shName == "common":
            continue
        flNames = [shName + ".html", "scripts/" + shName + "-markup.js"]
        for srcFile in flNames:
            src_path = src + "/" + srcFile
            if not os.path.exists(src_path):
                print("File not exists: " + src_path)
                continue
            destFile = dst + "/{lang}/" + srcFile
            sheet = wb[shName]
            flText = readFile(src_path)
            print("Translate: " + src_path)
            fileTranslate(flText, sheet, commonSheet, destFile)

def main():
    dstPath = "../langs"
    tmpPath = "../../langs"
    
    if os.path.exists(dstPath):
        shutil.rmtree(dstPath)
    wb = load_workbook('./dictionary.xlsx')
    translateAll(wb, "..", tmpPath)
    shutil.move(tmpPath, dstPath)
    print("Done!")
    
main()
